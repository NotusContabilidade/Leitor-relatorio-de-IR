# --------------------------------------------------------------------------
# ARQUIVO: app.py - Leitor de Relatório de IR
# --------------------------------------------------------------------------

import eel
import os
import base64
import io
import zipfile
import traceback
import concurrent.futures
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import sqlite3
import json
from datetime import datetime
import uuid

import engine
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Design Nótus ---
NOTUS_RED_HEX = "923240"
HEADER_FILL = PatternFill(start_color=NOTUS_RED_HEX, end_color=NOTUS_RED_HEX, fill_type="solid")
TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DATA_FONT = Font(name="Calibri", size=10)
BORDER_FULL = Border(
    left=Side(border_style="thin", color="D1D5DB"),
    right=Side(border_style="thin", color="D1D5DB"),
    top=Side(border_style="thin", color="D1D5DB"),
    bottom=Side(border_style="thin", color="D1D5DB")
)

eel.init('web')
DB_NAME = 'notus_ir_historico.db'

def _init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS historico_analises (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sessao_id TEXT, data_analise TEXT,
            cpf_cnpj TEXT, nome TEXT, status TEXT,
            detalhes_json TEXT, arquivo_pdf TEXT
        )
    """)
    conn.commit()
    conn.close()

def _salvar_historico(dados_json):
    sessao_id = str(uuid.uuid4())
    data_analise = datetime.now().isoformat()
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        for item in dados_json:
            if not item.get('nome'):
                continue
            c.execute("""
                INSERT INTO historico_analises
                (sessao_id, data_analise, cpf_cnpj, nome, status, detalhes_json, arquivo_pdf)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                sessao_id, data_analise,
                item.get('cpf', ''), item.get('nome', ''),
                item.get('status', ''),
                json.dumps(item), item.get('arquivo_pdf', '')
            ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[ERRO DB] {e}")

def aplicar_estetica_premium(ws, subtitulo):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    top_cell = ws.cell(row=1, column=1)
    top_cell.value = f"LEITOR DE RELATÓRIO DE IR - NÓTUS CONTÁBIL | {subtitulo}"
    top_cell.font = TITLE_FONT
    top_cell.fill = HEADER_FILL
    top_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = [str(ws.cell(row=2, column=c).value).upper() for c in range(1, ws.max_column + 1)]

    for cell in ws[2]:
        cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_FULL

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, cell in enumerate(row, start=1):
            cell.font = DATA_FONT
            cell.border = BORDER_FULL
            h_name = headers[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                if "VALOR" in h_name or "TOTAL" in h_name or "R$" in h_name:
                    cell.number_format = '"R$ "#,##0.00'
                    cell.alignment = Alignment(horizontal="right", wrap_text=False)
                else:
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=False, shrink_to_fit=False)

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if len(str(cell.value or "")) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = min(max_length + 5, 80)

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A3"

@eel.expose
def carregar_historico():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT sessao_id, data_analise, COUNT(*) as total
            FROM historico_analises
            GROUP BY sessao_id
            ORDER BY data_analise DESC
            LIMIT 50
        """)
        rows = [{"sessao_id": r[0], "data_analise": r[1], "total": r[2]} for r in c.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        print(f"[ERRO DB] {e}")
        return []

@eel.expose
def salvar_excel(dados_json, nome_sugerido="Relatorio"):
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        nome_final = f"Relatorio_IR_Notus_{os.path.splitext(nome_sugerido)[0]}.xlsx"
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=nome_final)
        root.destroy()
        if not caminho:
            return {"sucesso": False, "erro": "Cancelado"}

        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
        resumo = []
        detalhe = []

        for item in dados_json:
            resumo.append(engine.montar_linha_resumo(item))
            for linha in engine.montar_linhas_detalhe(item):
                detalhe.append(linha)

        div  = []
        isen = []
        excl = []
        rv   = []
        for item in dados_json:
            div.extend(engine.montar_linhas_dividas(item))
            isen.extend(engine.montar_linhas_rendimentos(item, 'rendimentos_isentos'))
            excl.extend(engine.montar_linhas_rendimentos(item, 'rendimentos_exclusivos'))
            rv.extend(engine.montar_linhas_renda_variavel(item))

        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            pd.DataFrame(resumo).to_excel(writer, sheet_name='Resumo', index=False)
            aplicar_estetica_premium(writer.sheets['Resumo'], f"RESUMO EM {agora}")

            if detalhe:
                pd.DataFrame(detalhe).to_excel(writer, sheet_name='BENS E DIREITOS', index=False)
                aplicar_estetica_premium(writer.sheets['BENS E DIREITOS'], "BENS E DIREITOS")

            if div:
                pd.DataFrame(div).to_excel(writer, sheet_name='DÍVIDAS E ÔNUS REAIS', index=False)
                aplicar_estetica_premium(writer.sheets['DÍVIDAS E ÔNUS REAIS'], "DÍVIDAS E ÔNUS REAIS")

            if isen:
                pd.DataFrame(isen).to_excel(writer, sheet_name='RENDIMENTOS ISENTOS', index=False)
                aplicar_estetica_premium(writer.sheets['RENDIMENTOS ISENTOS'], "RENDIMENTOS ISENTOS")

            if excl:
                pd.DataFrame(excl).to_excel(writer, sheet_name='REND. EXCLUSIVA', index=False)
                aplicar_estetica_premium(writer.sheets['REND. EXCLUSIVA'],
                                         "RENDIMENTOS SUJEITOS À TRIBUTAÇÃO EXCLUSIVA")

            if rv:
                pd.DataFrame(rv).to_excel(writer, sheet_name='RENDA VARIÁVEL', index=False)
                aplicar_estetica_premium(writer.sheets['RENDA VARIÁVEL'],
                                         "RENDA VARIÁVEL - DAY-TRADE E FII")

        return {"sucesso": True, "caminho": caminho}
    except Exception as e:
        return {"sucesso": False, "erro": str(e)}

# --- Funções de Processamento ---

def _worker_zip(args):
    bz, pasta, nome_pdf = args
    with zipfile.ZipFile(io.BytesIO(bz), 'r') as zf:
        return engine.analisar_pdf_worker(zf.read(nome_pdf), pasta, os.path.basename(nome_pdf))

@eel.expose
def processar_pdfs_soltos_paralelo(payloads):
    tarefas = [(base64.b64decode(p['bytes']), "Upload", p['nome']) for p in payloads]
    res = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as exc:
        futures = [exc.submit(lambda x: engine.analisar_pdf_worker(*x), t) for t in tarefas]
        for i, f in enumerate(concurrent.futures.as_completed(futures)):
            res.append(f.result())
            eel.update_progress(int(((i + 1) / len(tarefas)) * 100), f"Lendo PDFs: {i+1}/{len(tarefas)}...")()
    _salvar_historico(res)
    return {"sucesso": True, "dados": res}

@eel.expose
def processar_arquivo_zip_paralelo(file_base64, nome_arquivo):
    bz = base64.b64decode(file_base64)
    tarefas = []
    with zipfile.ZipFile(io.BytesIO(bz), 'r') as zf:
        for info in zf.infolist():
            if info.is_dir() or not info.filename.lower().endswith('.pdf'):
                continue
            tarefas.append((bz, os.path.dirname(info.filename), info.filename))
    res = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as exc:
        futures = [exc.submit(_worker_zip, t) for t in tarefas]
        for i, f in enumerate(concurrent.futures.as_completed(futures)):
            res.append(f.result())
            eel.update_progress(int(((i + 1) / len(tarefas)) * 100), f"Processando ZIP: {i+1}/{len(tarefas)}...")()
    _salvar_historico(res)
    return {"sucesso": True, "dados": res}

if __name__ == "__main__":
    _init_db()
    try:
        eel.start('index.html', size=(1400, 900))
    except:
        pass
